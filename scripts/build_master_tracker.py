#!/usr/bin/env python3

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent.parent
OUTPUT = ROOT / "docs" / "tracking" / "ailit_master_tracker.xlsx"


BACKLOG_ROWS = [
    [
        "BK-001",
        "Planning",
        "Requirements",
        "Create compounding knowledge requirements document",
        "Capture the problem, scope, and success criteria before implementation.",
        "P0",
        "Done",
        "Phase 0",
        "Org",
        "",
        "docs/brainstorms/ailit-compounding-knowledge-requirements.md",
        "Platform",
        "Completed on 2026-05-06.",
    ],
    [
        "BK-002",
        "Planning",
        "Implementation Plan",
        "Create compounding knowledge implementation plan",
        "Define the phased delivery path and parallel worktree split.",
        "P0",
        "Done",
        "Phase 0",
        "Org",
        "BK-001",
        "docs/plans/ailit-compounding-knowledge-plan.md",
        "Platform",
        "Completed on 2026-05-06.",
    ],
    [
        "BK-003",
        "Foundation",
        "Artifact Model",
        "Define solution metadata contract",
        "Create a stable structure for reusable solved-problem docs.",
        "P0",
        "Done",
        "Phase 1",
        "Org",
        "BK-002",
        "docs/solutions/_template.md",
        "Docs lane",
        "Completed in the first artifact-model batch.",
    ],
    [
        "BK-004",
        "Foundation",
        "Artifact Model",
        "Create shared solution template and README",
        "Give engineers a simple, repeatable way to store reusable solutions.",
        "P0",
        "Done",
        "Phase 1",
        "Org",
        "BK-003",
        "docs/solutions/README.md",
        "Docs lane",
        "Completed in the first artifact-model batch.",
    ],
    [
        "BK-005",
        "Foundation",
        "Documentation",
        "Update docs to explain knowledge layers",
        "Prevent confusion between standards, skills, solutions, decisions, and lessons.",
        "P1",
        "Done",
        "Phase 1",
        "Org",
        "BK-003",
        "docs/skills.md",
        "Docs lane",
        "Overview and skills docs updated in the first artifact-model batch.",
    ],
    [
        "BK-006",
        "Registry",
        "Solution Discovery",
        "Add solution registry builder",
        "Make solution docs searchable and loadable by the runtime.",
        "P0",
        "Done",
        "Phase 2",
        "Org",
        "BK-004",
        "aios/solution_registry.py",
        "Registry lane",
        "Implemented with registry/solutions.json output.",
    ],
    [
        "BK-007",
        "Registry",
        "Validation",
        "Validate solution metadata and file shape",
        "Keep shared knowledge high quality and predictable.",
        "P0",
        "Done",
        "Phase 2",
        "Org",
        "BK-006",
        "registry/solutions.json",
        "Registry lane",
        "Validation now blocks invalid solution indexing.",
    ],
    [
        "BK-008",
        "Registry",
        "CLI",
        "Add list-solutions command",
        "Let engineers inspect which reusable solutions are available.",
        "P1",
        "Done",
        "Phase 2",
        "Team",
        "BK-006",
        "aios/cli.py",
        "Registry lane",
        "Implemented with plain output and JSON mode.",
    ],
    [
        "BK-009",
        "Registry",
        "Search",
        "Add search/filter support for solutions",
        "Match solution docs by stack tags and problem keywords.",
        "P1",
        "Done",
        "Phase 2",
        "Team",
        "BK-006",
        "aios/solution_registry.py",
        "Registry lane",
        "Implemented with query, status, and trust filters.",
    ],
    [
        "BK-010",
        "Capture",
        "Knowledge Capture",
        "Add capture-solution command",
        "Turn solved work into a structured reusable solution note.",
        "P0",
        "Done",
        "Phase 3",
        "Team",
        "BK-004, BK-006",
        "aios/capture_solution.py",
        "Runtime lane",
        "Implemented via CLI and shared solution document writer.",
    ],
    [
        "BK-011",
        "Capture",
        "Promotion",
        "Add promote-lesson command",
        "Move good project lessons into shared reusable knowledge.",
        "P0",
        "Done",
        "Phase 3",
        "Team",
        "BK-010",
        "aios/promote_lesson.py",
        "Runtime lane",
        "Implemented with lesson parsing from ai/lessons.md.",
    ],
    [
        "BK-012",
        "Capture",
        "CLI",
        "Add list-knowledge command",
        "Give one place to inspect skills, solutions, and local knowledge artifacts.",
        "P1",
        "Done",
        "Phase 3",
        "Team",
        "BK-006, BK-010",
        "aios/cli.py",
        "Runtime lane",
        "Implemented with optional project-local sections and JSON output.",
    ],
    [
        "BK-013",
        "Runtime",
        "Prepare",
        "Load matching solutions during prepare",
        "Let the agent reuse proven solved patterns during task preparation.",
        "P0",
        "Done",
        "Phase 4",
        "Project",
        "BK-006, BK-009",
        "aios/prepare.py",
        "Runtime lane",
        "Implemented with solution matching and context sections in prepare.",
    ],
    [
        "BK-014",
        "Runtime",
        "Prompt Quality",
        "Bound solution context size in prepare",
        "Avoid token explosion when many matches exist.",
        "P0",
        "Done",
        "Phase 4",
        "Project",
        "BK-013",
        "aios/context_builder.py",
        "Runtime lane",
        "Implemented with solution limits and excerpt truncation.",
    ],
    [
        "BK-015",
        "Runtime",
        "Observability",
        "Add runtime audit logging",
        "Show whether AIlit was actually used during real tasks.",
        "P1",
        "Done",
        "Phase 4",
        "Org",
        "BK-013",
        "aios/prepare.py",
        "Runtime lane",
        "Implemented with ai/usage.log entries from prepare.",
    ],
    [
        "BK-016",
        "Governance",
        "Quality",
        "Add trust and review status for shared knowledge",
        "Shared knowledge should show whether it is draft or approved.",
        "P1",
        "Done",
        "Phase 5",
        "Org",
        "BK-003, BK-006",
        "registry/solutions.json",
        "Governance lane",
        "Implemented with trust, review_status, reviewed_by, and last_reviewed.",
    ],
    [
        "BK-017",
        "Governance",
        "Freshness",
        "Detect stale or unreviewed solutions",
        "Keep shared knowledge from decaying silently.",
        "P2",
        "Done",
        "Phase 5",
        "Org",
        "BK-016",
        "aios/solution_registry.py",
        "Governance lane",
        "Implemented with age_days, is_stale, and list filters.",
    ],
    [
        "BK-018",
        "Testing",
        "Registry Tests",
        "Add tests for solution indexing and validation",
        "Prevent registry regressions as the knowledge model grows.",
        "P0",
        "Planned",
        "Phase 2",
        "Org",
        "BK-006, BK-007",
        "tests/",
        "QA lane",
        "Need fixture solution docs.",
    ],
    [
        "BK-019",
        "Testing",
        "Capture Tests",
        "Add tests for capture and promotion commands",
        "Keep post-work capture safe and predictable.",
        "P0",
        "Planned",
        "Phase 3",
        "Team",
        "BK-010, BK-011",
        "tests/",
        "QA lane",
        "Verify metadata mapping and filesystem output.",
    ],
    [
        "BK-020",
        "Testing",
        "Runtime Tests",
        "Add tests for prepare integration with solutions",
        "Ensure runtime retrieval stays clean and bounded.",
        "P0",
        "Planned",
        "Phase 4",
        "Project",
        "BK-013, BK-014",
        "tests/",
        "QA lane",
        "Verify relevance ranking and output sections.",
    ],
    [
        "BK-021",
        "Rollout",
        "Org Setup",
        "Document shared team or org rollout model",
        "Explain how to run one shared AIlit repo across many engineers and repos.",
        "P1",
        "Done",
        "Phase 5",
        "Org",
        "BK-005",
        "docs/team_setup.md",
        "Docs lane",
        "Added team setup and governance guidance.",
    ],
]


ARTIFACT_ROWS = [
    ["Standards", "Stable reusable rules", "Global or org", "Engineering principles and approved rules"],
    ["Skills", "Reusable guidance modules", "Global, team, or imported", "How to approach a class of work"],
    ["Solutions", "Solved problem writeups", "Shared team or org", "How a real problem was solved and when to reuse it"],
    ["Decisions", "Project-local decision memory", "Per project", "Why a project chose a specific direction"],
    ["Lessons", "Project-local learnings", "Per project", "What was learned from recent work or incidents"],
]


OPEN_QUESTIONS = [
    ["Q-001", "Where should reviewed shared solutions live long term?", "Open", "Use docs/solutions/ in the first pass.", "Platform"],
    ["Q-002", "How much of promotion should be automated?", "Open", "Start manual with good scaffolding and suggestions.", "Platform"],
    ["Q-003", "What is the minimum trust metadata for shared knowledge?", "Open", "Owner, source, review status, trust level, updated date.", "Platform"],
    ["Q-004", "How should stack-specific reusable patterns be tagged?", "Open", "Use simple tags first: language, framework, domain, infra.", "Platform"],
]


FILL = {
    "navy": PatternFill("solid", fgColor="1F4E78"),
    "blue": PatternFill("solid", fgColor="D9EAF7"),
    "green": PatternFill("solid", fgColor="E2F0D9"),
    "amber": PatternFill("solid", fgColor="FFF2CC"),
    "gray": PatternFill("solid", fgColor="F3F4F6"),
    "done": PatternFill("solid", fgColor="D9EAD3"),
    "planned": PatternFill("solid", fgColor="FFF2CC"),
    "in_progress": PatternFill("solid", fgColor="D9EAF7"),
}

THIN = Side(style="thin", color="D0D7DE")


def style_header(row):
    for cell in row:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = FILL["navy"]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_grid(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autofit(ws, widths):
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_summary(ws):
    ws.title = "Summary"
    ws["A1"] = "AIlit Master Build Tracker"
    ws["A1"].font = Font(size=16, bold=True, color="1F1F1F")
    ws["A2"] = "Purpose"
    ws["B2"] = "Track the compounding knowledge and maintenance backlog before splitting implementation across worktrees."
    ws["A4"] = "Current planning artifacts"
    ws["A5"] = "Requirements doc"
    ws["B5"] = "docs/brainstorms/ailit-compounding-knowledge-requirements.md"
    ws["A6"] = "Implementation plan"
    ws["B6"] = "docs/plans/ailit-compounding-knowledge-plan.md"
    ws["A8"] = "KPI"
    ws["B8"] = "Value"
    style_header(ws[8])
    metrics = [
        ("Total backlog items", '=COUNTA(Backlog!A2:A200)'),
        ("Done", '=COUNTIF(Backlog!G2:G200,"Done")'),
        ("In progress", '=COUNTIF(Backlog!G2:G200,"In Progress")'),
        ("Planned", '=COUNTIF(Backlog!G2:G200,"Planned")'),
        ("P0 items", '=COUNTIF(Backlog!F2:F200,"P0")'),
        ("Org-level items", '=COUNTIF(Backlog!I2:I200,"Org")'),
    ]
    row = 9
    for label, formula in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=formula)
        row += 1
    ws["D8"] = "Phase"
    ws["E8"] = "Focus"
    style_header(ws[8][3:5])
    phase_rows = [
        ("Phase 1", "Artifact model, templates, docs"),
        ("Phase 2", "Solution registry and validation"),
        ("Phase 3", "Capture and promotion commands"),
        ("Phase 4", "Prepare integration and auditability"),
        ("Phase 5", "Governance, freshness, org rollout"),
    ]
    row = 9
    for phase, focus in phase_rows:
        ws.cell(row=row, column=4, value=phase)
        ws.cell(row=row, column=5, value=focus)
        row += 1
    ws["D16"] = "Execution note"
    ws["E16"] = "Use the backlog sheet as the working queue. Split lanes by docs, registry, runtime, governance, and QA."
    for cell in ["A2", "A4", "A8", "D8", "D16"]:
        ws[cell].font = Font(bold=True)
    style_grid(ws)
    autofit(ws, {1: 28, 2: 70, 4: 18, 5: 48})
    ws.freeze_panes = "A8"


def build_backlog(ws):
    ws.title = "Backlog"
    headers = [
        "ID",
        "Epic",
        "Workstream",
        "Task",
        "Why It Matters",
        "Priority",
        "Status",
        "Phase",
        "Knowledge Level",
        "Depends On",
        "Target Artifact",
        "Suggested Owner",
        "Notes",
    ]
    ws.append(headers)
    for row in BACKLOG_ROWS:
        ws.append(row)
    style_header(ws[1])
    style_grid(ws)
    for r in range(2, ws.max_row + 1):
        status = ws.cell(r, 7).value
        fill = None
        if status == "Done":
            fill = FILL["done"]
        elif status == "In Progress":
            fill = FILL["in_progress"]
        elif status == "Planned":
            fill = FILL["planned"]
        if fill:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = fill
    table = Table(displayName="BacklogTable", ref=f"A1:M{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    autofit(
        ws,
        {
            1: 10,
            2: 16,
            3: 18,
            4: 42,
            5: 44,
            6: 10,
            7: 14,
            8: 12,
            9: 16,
            10: 16,
            11: 40,
            12: 16,
            13: 48,
        },
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{ws.max_row}"


def build_artifacts(ws):
    ws.title = "Knowledge Model"
    ws.append(["Artifact Type", "Purpose", "Scope", "Example Use"])
    for row in ARTIFACT_ROWS:
        ws.append(row)
    style_header(ws[1])
    style_grid(ws)
    ws["A8"] = "Design rule"
    ws["B8"] = "Do not force solved-problem writeups into skills. Solutions should remain a separate reusable knowledge type."
    ws["A8"].font = Font(bold=True)
    ws["A8"].fill = FILL["blue"]
    ws["B8"].fill = FILL["blue"]
    autofit(ws, {1: 18, 2: 32, 3: 22, 4: 56})
    ws.freeze_panes = "A2"


def build_questions(ws):
    ws.title = "Open Questions"
    ws.append(["ID", "Question", "Status", "Current Direction", "Owner"])
    for row in OPEN_QUESTIONS:
        ws.append(row)
    style_header(ws[1])
    style_grid(ws)
    autofit(ws, {1: 10, 2: 54, 3: 12, 4: 44, 5: 14})
    ws.freeze_panes = "A2"


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_summary(wb.active)
    build_backlog(wb.create_sheet())
    build_artifacts(wb.create_sheet())
    build_questions(wb.create_sheet())
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
